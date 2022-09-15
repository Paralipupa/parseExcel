import abc
import os

from openpyxl import load_workbook
import xlrd


def rchop(s, sub):
    return s[:-len(sub)] if s.endswith(sub) else s


class DataFile(abc.ABC):

    def __init__(self, fname: str, sheet_name: str, columns: range):
        self.fname = fname
        self.sheet_name = sheet_name
        self.columns = columns

    def __iter__(self):
        return self

    def __next__(self):
        return ""


class XlsFile(DataFile):
    def __init__(self, fname: str, sheet_name: str = None, columns: range = range(20), page_index: int = 0):
        super(XlsFile, self).__init__(fname, sheet_name, columns)
        self._book = xlrd.open_workbook(fname, logfile=open(os.devnull, 'w'))
        if self.sheet_name:
            sheet = self._book.sheet_by_name(self.sheet_name)
        else:
            sheet = self._book.sheets()[page_index]
        self._rows = (sheet.row(index) for index in range(sheet.nrows))

    @staticmethod
    def get_cell_text(cell):
        if cell.ctype == 2:
            return rchop(str(cell.value), '.0')
        return str(cell.value)

    def get_row(self, row):
        index = 0
        for cell in row:
            if index in self.columns:
                yield XlsFile.get_cell_text(cell)
            index = index + 1

    def __next__(self):
        for row in self._rows:
            return list(self.get_row(row))
        raise StopIteration

    def __del__(self):
        pass


class XlsxFile(DataFile):
    def __init__(self, fname: str, sheet_name: str = None, columns: range = range(20), page_index: int = 0):
        super(XlsFile, self).__init__(fname, sheet_name, columns)
        self._wb = load_workbook(filename=fname, read_only=True)
        if self.sheet_name:
            self._ws = self._wb.get_sheet_by_name(self.sheet_name)
        else:
            self._ws = self._wb.worksheets[page_index]
        self._cursor = self._ws.iter_rows()
        row_num = 0
        while row_num < self._first_line:
            row_num += 1
            next(self._cursor)

    @staticmethod
    def get_cell_text(cell):
        return str(cell.value) if cell.value else ""

    def get_row(self, row):
        i = 0
        for cell in row:
            if i in self.columns:
                yield XlsxFile.get_cell_text(cell)
            i += 1

    def __next__(self):
        return list(self.get_row(next(self._cursor)))

    def __del__(self):
        self._wb.close()

    def get_index(self, cell):
        try:
            return cell.column
        except AttributeError:
            return -1


def get_file_reader(fname):
    """Get class for reading file as iterable"""
    _, file_extension = os.path.splitext(fname)
    if file_extension == '.xls':
        return XlsFile
    if file_extension == '.xlsx':
        return XlsxFile
    raise Exception("Unknown file type")
